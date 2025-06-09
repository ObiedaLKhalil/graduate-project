from multiprocessing import Process, Value
from threading import Thread
import time
from flask import Flask, request, jsonify, send_from_directory,send_file
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import json
import firebase_admin
from firebase_admin import credentials, messaging
from flask_cors import CORS
import requests
import shutil
import lgpio
from picamera2 import Picamera2
from openpyxl import load_workbook
import cv2
import numpy as np
from PIL import Image
import tempfile
from deepface import DeepFace
from sklearn.preprocessing import normalize
import albumentations as A
import random
import joblib
from openpyxl import Workbook
from datetime import datetime
import openpyxl

# Flask App
app = Flask(__name__)
CORS(app)  # Enable CORS to allow requests from the React Native app

# Load Firebase credentials (Do this ONCE, outside __main__)
cred = credentials.Certificate("/home/obied/smart-attendance-system-19896-firebase-adminsdk-fbsvc-a5c987922b.json")
firebase_admin.initialize_app(cred)

# Device token holder
device_token = ""

# Directory to save uploaded images
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'downloads')
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)


up_path = '/home/obied/up'
os.makedirs(up_path, exist_ok=True)


down_path = '/home/obied/down'
os.makedirs(down_path, exist_ok=True)

# Path to save captured images
IMAGE_DIR = '/home/obied/captured_images'
if not os.path.exists(IMAGE_DIR):
    os.makedirs(IMAGE_DIR)


UOUTPUT_FOLDER = os.path.join(os.getcwd(), 'output')
os.makedirs(UOUTPUT_FOLDER, exist_ok=True)

clf = joblib.load("/home/obied/face_recognizer.pkl")
label_map = joblib.load("/home/obied/label_map.pkl")
faces = joblib.load("/home/obied/faces_vectors.pkl")
labels = joblib.load("/home/obied/labels_array.pkl")


pir1 = 17        # GPIO17 = Physical Pin 11
pir2 = 27        # GPIO27 = Physical Pin 13
TRIG = 23        # GPIO23 = Physical Pin 16
ECHO = 24        # GPIO24 = Physical Pin 18
relay_pin = 26   # GPIO22 = Physical Pin 15
h = lgpio.gpiochip_open(0)
lgpio.gpio_claim_input(h, pir1)
lgpio.gpio_claim_input(h, pir2)
lgpio.gpio_claim_input(h, ECHO)
lgpio.gpio_claim_output(h, relay_pin)
lgpio.gpio_claim_output(h, TRIG)

picam0 = Picamera2()
picam1 = Picamera2(camera_num=1)

# Configure the camera for still image capture
picam0.configure(picam0.create_still_configuration(main={"size": ( 2592, 1944 )}))
picam1.configure(picam1.create_still_configuration(main={"size": (2048, 1536 )}))  # IMX219 max res

is_training = Value('b', False)
training_condition=False
entry_cond=False
exit_cond=False
def capture_image1():
    # Start the camera
    picam0.start()
    time.sleep(2)  # Let camera adjust
    # Capture image
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    image_path_0 = os.path.join(IMAGE_DIR, f"image_cam0_{timestamp}.jpg")

    # Capture and save the image
    picam0.capture_file(image_path_0)

    # Stop the camera after capture
    picam0.stop()

    return image_path_0

def capture_image2():

    # Start the camera
    picam1.start()
    time.sleep(2)  # Let camera adjust

    # Create timestamped filename
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    image_path_1 = os.path.join(IMAGE_DIR, f"image_cam1_{timestamp}.jpg")

    # Capture and save the image
    picam1.capture_file(image_path_1)
    print(image_path_1)
    # Stop the camera after capture
    picam1.stop()

    return image_path_1

def delete_all_files_in_folder(folder_path):
    try:
        # Loop through all items in the folder
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            # Only delete files (not subdirectories)
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"Deleted file: {file_path}")
        print("All files have been deleted.")
    except Exception as e:
        print(f"Error: {e}")


def get_distance():

    # Send 10us pulse to trigger
    lgpio.gpio_write(h, TRIG, 1)
    time.sleep(0.00001) # 10 microseconds
    lgpio.gpio_write(h, TRIG, 0)

     # Wait for echo start
    start_time = time.time()
    while lgpio.gpio_read(h, ECHO) == 0:
        start_time = time.time()

    # Wait for echo end
    stop_time = time.time()
    while lgpio.gpio_read(h, ECHO) == 1:
        stop_time = time.time()

    # Calculate distance
    time_elapsed = stop_time - start_time
    distance = (time_elapsed * 34300) / 2 # Speed of sound: 34300 cm/s
    return distance


def recognize_face(face_image, threshold=0.5):
    global clf, label_map,faces,labels
    try:
        # Extract embedding
        embedding = DeepFace.represent(img_path=face_image, model_name="Facenet", enforce_detection=False)
        face_embedding = np.array(embedding[0]["embedding"])
        face_embedding = normalize(face_embedding.reshape(1, -1))[0]

        # Get prediction and probability
        label = clf.predict([face_embedding])[0]
        confidence = np.max(clf.predict_proba([face_embedding]))

        # Compute similarity with all known faces
        similarities = np.dot(faces, face_embedding.T)  # Cosine similarity
        max_similarity = np.max(similarities)

        if max_similarity > threshold:  # Ensure it's close to known faces
            print(f"Recognized as: {label_map[label]} with confidence: {confidence:.2f}")
            return 1, label_map[label]
        else:
            print("Unknown person (face not in dataset)")
            return 0, None
    except Exception as e:
        print(f"Error recognizing face: {e}")
def detect_faces_and_get_images(image_path):
    net = cv2.dnn.readNetFromCaffe("/home/obied/deploy.prototxt", "/home/obied/res10_300x300_ssd_iter_140000.caffemodel")
    save_directory = "cropped_faces"  # Directory where cropped images will be saved

# Ensure the save directory exists
    if not os.path.exists(save_directory):
      os.makedirs(save_directory)
    # Load the image
    pil_image = Image.open(image_path)
    frame = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
    h, w = frame.shape[:2]
    print("Image loaded successfully.")
    # Perform face detection
    blob = cv2.dnn.blobFromImage(frame, 1.0, (300, 300), (104.0, 177.0, 123.0))
    net.setInput(blob)
    detections = net.forward()
    detected_faces_paths=[]
    faces = []  # List to hold cropped faces
    face_count = 0  # Counter for saving faces
    # Assuming w and h are the width and height of the image, and detections is the output from the model.
    for i in range(detections.shape[2]):
     confidence = detections[0, 0, i, 2]
     if confidence > 0.5:  # Adjust confidence threshold
        # Get the original bounding box coordinates
        box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
        (startX, startY, endX, endY) = box.astype("int")

        # Define a scaling factor (e.g., 1.2 will increase the bounding box by 20%)
        scale_factor = 1.2

        # Calculate the width and height of the bounding box
        width = endX - startX
        height = endY - startY

        # Increase the width and height by the scale factor
        width = int(width * scale_factor)
        height = int(height * scale_factor)

        # Update the start and end points of the bounding box to enlarge it
        # Adjust the start and end points to keep the bounding box centered
        startX = max(0, startX - (width - (endX - startX)) // 2)  # Ensure the box stays within the image
        startY = max(0, startY - (height - (endY - startY)) // 2)
        endX = min(w, startX + width)  # Ensure the box stays within the image bounds
        endY = min(h, startY + height)
        # Crop the face from the image using the new bounding box
        face = frame[startY:endY, startX:endX]
        resized_face = cv2.resize(face, (160, 160))
        normalized_face = resized_face.astype("float32") / 255.0
        input_face = np.expand_dims(normalized_face, axis=0)
        save_path = os.path.join(save_directory,f"cropped_face{face_count}.jpg")
        cv2.imwrite(save_path, (input_face[0] * 255).astype(np.uint8))
        if os.path.exists(save_path):
          print(f"Face {face_count} saved successfully at {save_path}")
        else:
          print(f"Error saving face {face_count}")

        detected_faces_paths.append(save_path)  # Store file path
        faces.append(face)  # Append the enlarged face to the list
        face_count += 1
    return detected_faces_paths  # Return the list of faces


def gamma_correction(img, gamma=1.5):
    invGamma = 1.0 / gamma
    table = np.array([(i / 255.0) ** invGamma * 255 for i in range(256)]).astype("uint8")
    return cv2.LUT(img, table)

def histogram_equalization(img):
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    return clahe.apply(img)

def add_gaussian_noise(img):
    row, col = img.shape
    mean, sigma = 0, 15
    gauss = np.random.normal(mean, sigma, (row, col)).astype("uint8")
    return cv2.add(img, gauss)

def add_salt_pepper_noise(img, salt_prob=0.02, pepper_prob=0.02):
    noisy_img = img.copy()
    total_pixels = img.size
    num_salt = int(total_pixels * salt_prob)
    num_pepper = int(total_pixels * pepper_prob)

    coords = [np.random.randint(0, i, num_salt) for i in img.shape]
    noisy_img[coords[0], coords[1]] = 255

    coords = [np.random.randint(0, i, num_pepper) for i in img.shape]
    noisy_img[coords[0], coords[1]] = 0

    return noisy_img

def apply_affine_transform(img):
    rows, cols = img.shape
    src_pts = np.float32([[0, 0], [cols-1, 0], [0, rows-1]])
    dst_pts = src_pts + np.random.randint(-10, 10, size=src_pts.shape).astype(np.float32)
    M = cv2.getAffineTransform(src_pts, dst_pts)
    return cv2.warpAffine(img, M, (cols, rows))

def apply_random_occlusion(img):
    occluded_img = img.copy()
    h, w = img.shape
    x, y = random.randint(0, w // 2), random.randint(0, h // 2)
    x2, y2 = x + random.randint(20, 50), y + random.randint(20, 50)
    occluded_img[y:y2, x:x2] = 0
    return occluded_img

def flip_image(img):
    return cv2.flip(img, 1)

def rotate_image(img, angle=10):
    h, w = img.shape
    M = cv2.getRotationMatrix2D((w // 2, h // 2), angle, 1)
    return cv2.warpAffine(img, M, (w, h))
def mask_half_face(img, side="right"):
    """
    Masks one half of the face to simulate a partial view.
    :param img: Input image.
    :param side: Choose either "left" or "right" to mask half.
    :return: Augmented image with half of the face masked.
    """
    h, w = img.shape
    if side == "left":
        img[:, :w//2] = 0  # Set left half to black
    else:
        img[:, w//2:] = 0  # Set right half to black
    return img

augmentations = [
    gamma_correction, histogram_equalization, add_gaussian_noise,
    add_salt_pepper_noise, apply_affine_transform, apply_random_occlusion,
    flip_image, rotate_image,mask_half_face
]

def augment_images_in_image_set():
    # List of files in imageSet folder
    files_in_image_set = os.listdir(DOWNLOAD_FOLDER)

    # If there are files to process
    if files_in_image_set:
        for image_name in files_in_image_set:
            image_path = os.path.join(DOWNLOAD_FOLDER, image_name)
            folder_name = os.path.splitext(image_name)[0]  # Get the name without the extension
            image_output_folder = os.path.join(UOUTPUT_FOLDER, folder_name)
            os.makedirs(image_output_folder, exist_ok=True)

            # Check if it is an image
            if image_name.lower().endswith(('.jpg', '.jpeg', '.png')):
                image = cv2.imread(image_path)
                grayimage = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # Convert to grayscale for LBP

                if image is not None:
                    # Generate 300 augmented images for each image
                    for i in range(300):
                        aug_image = grayimage.copy()
                        num_augmentations = random.randint(1, 3)
                        chosen_augmentations = random.sample(augmentations, num_augmentations)
                        for aug in chosen_augmentations:
                            aug_image = aug(aug_image)

                        output_path = os.path.join(image_output_folder, f"augmented_{i+1}_{image_name}")
                        cv2.imwrite(output_path, aug_image)

                    print(f"Augmented 300 images for: {image_name}")
                else:
                    print(f"Failed to load image: {image_name}")
            else:
                print(f"Skipping non-image file: {image_name}")


def update_model_with_new_data(new_dataset_path):
    global clf, label_map

    # Load old embeddings, labels, and label_map
    try:
        faces = joblib.load("/home/obied/faces_vectors.pkl")
        labels = joblib.load("/home/obied/labels_array.pkl")
        print("✅ Loaded previous data.")
    except:
        faces = []
        labels = []
        label_map = {}
        print("⚠️ No previous model found, starting fresh.")

    current_label_index = max(label_map.keys()) + 1 if label_map else 0

    new_faces = []
    new_labels = []

    for person in os.listdir(new_dataset_path):
        person_path = os.path.join(new_dataset_path, person)

        # Skip non-directory files
        if not os.path.isdir(person_path):
            continue

        # If person already exists, reuse label
        if person in label_map.values():
            label = [k for k, v in label_map.items() if v == person][0]
        else:
            label = current_label_index
            label_map[label] = person
            current_label_index += 1

        for image_name in os.listdir(person_path):
            image_path = os.path.join(person_path, image_name)
            try:
                print("embedding")
                embedding = DeepFace.represent(img_path=image_path, model_name="Facenet", enforce_detection=False)
                print("embedding_vector")
                embedding_vector = np.array(embedding[0]["embedding"])
                print("embedding_vector")
                embedding_vector = normalize(embedding_vector.reshape(1, -1))[0]

                new_faces.append(embedding_vector)
                new_labels.append(label)
            except Exception as e:
                print(f"❌ Error processing {image_path}: {e}")

    if len(new_faces) == 0:
        print("⚠️ No new embeddings added. Model not updated.")
        return

    # Add new embeddings to old ones
    if len(faces) > 0:
        faces = np.vstack([faces, new_faces])
        labels = np.concatenate([labels, new_labels])
    else:
        faces = np.array(new_faces)
        labels = np.array(new_labels)

    # Check unique classes before training
    if len(set(labels)) <= 1:
        print("❌ Cannot train model: Not enough unique classes.")
        return

    # Retrain classifier
    clf = SVC(kernel='linear', probability=True)
    clf.fit(faces, labels)

    # Save updated model
    joblib.dump(clf, "/home/obied/face_recognizer.pkl")
    joblib.dump(label_map, "/home/obied/label_map.pkl")
    joblib.dump(faces, "/home/obied/faces_vectors.pkl")
    joblib.dump(labels, "/home/obied/labels_array.pkl")

    print("✅ Model successfully updated and saved.")


def train_model():
    global training_condition,is_training
    is_training.value = True
    print("[TRAINING] Model training started...")
    augment_images_in_image_set()
    update_model_with_new_data("/home/obied/output")
    delete_all_files_in_folder(DOWNLOAD_FOLDER)
    delete_all_files_in_folder(UOUTPUT_FOLDER)
    print("[TRAINING] Model training completed.")
    is_training.value = False

def record_attendance(recognized_label):
    already_checked_in = False  # Reset every time
    file_name = "/home/obied/attendance_data.xlsx"
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["employee name", "time of entry", "time of exit"])

    if entry_cond:
        name = recognized_label
        current_date1 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                exit_cell_value = sheet.cell(row=cell.row, column=3).value
                if cell.value == recognized_label :
                    if exit_cell_value is None or exit_cell_value == "":
                        already_checked_in = True
                        break

        if not already_checked_in:
            sheet.append([name, current_date1, ""])

    if exit_cond:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == recognized_label:
                    current_date2 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    sheet.cell(row=cell.row, column=3, value=current_date2)

    wb.save(file_name)
    print("Data recorded successfully.")


@app.route('/capture12/', methods=['POST'])
def captureFromBothCameras():
    try:
        # Capture images from both sources
        image_path_1 = capture_image2()
        image_path_0 = capture_image1()

        # Check if both images were successfully captured
        if not os.path.exists(image_path_1) or not os.path.exists(image_path_0):
            return jsonify({'status': 'One or both images not found'}), 404

        # Move the captured images to the upload folder
        for image_path in [image_path_0, image_path_1]:
            destination = os.path.join(UPLOAD_FOLDER, os.path.basename(image_path))
            shutil.move(image_path, destination)

        return jsonify({'status': 'Images captured and moved successfully'}), 200

    except Exception as e:
        return jsonify({'status': 'Error capturing images', 'error': str(e)}), 500


@app.route('/solenoid/<state>', methods=['POST'])
def control_solenoid(state):
    if state == 'on':
        # GPIO.output(SOLENOID_PIN, GPIO.HIGH)
        print("open")
        lgpio.gpio_write(h, relay_pin, 1)
        return jsonify({'status': 'Solenoid Opened'}), 200
    elif state == 'off':
        # GPIO.output(SOLENOID_PIN, GPIO.LOW)
        lgpio.gpio_write(h, relay_pin, 0)
        print("closed")
        return jsonify({'status': 'Solenoid Closed'}), 200
    else:
        return jsonify({'status': 'Invalid state'}), 400



# Path to your Excel file
EXCEL_FILE_PATH = "/home/obied/attendance_data.xlsx"
@app.route('/attendance', methods=['GET'])
def send_attendance_data():
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            print("✅ Excel file found.")
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]
            print("Headers:", headers)

            data_array = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for key, value in zip(headers, row):
                    # Convert time/datetime objects to strings
                    if hasattr(value, 'isoformat'):
                        row_dict[key] = value.isoformat()
                    else:
                        row_dict[key] = value
                data_array.append(row_dict)

            print("Data array:", data_array)

            return jsonify({"data": data_array})
        else:
            print("❌ Excel file not found.")
            return jsonify({"error": "Attendance file not found"}), 404

    except Exception as e:
        print("🔥 Error:", str(e))
        return jsonify({"error": str(e)}), 500



@app.route('/get-image', methods=['GET'])
def get_image():
    try:
        # Get all files in the uploads directory
        files = [f for f in os.listdir(UPLOAD_FOLDER) if os.path.isfile(os.path.join(UPLOAD_FOLDER, f))]

        # Filter only image files (optional, based on extensions)
        image_files = [f for f in files if f.lower().endswith(('.jpg', '.jpeg', '.png'))]

        if not image_files:
            return "No images left", 404

        # Pick the first image (you could sort if you want a specific order)
        image_name = image_files[0]
        image_path = os.path.join(UPLOAD_FOLDER, image_name)

        # Send the image
        response = send_file(image_path, mimetype='image/jpeg', as_attachment=False)

        # Delete the image AFTER sending (optional: use a safer approach in production)
        os.remove(image_path)

        return response

    except Exception as e:
        return str(e), 500

# Allowed file extensions for image upload
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# Check allowed file extensions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload', methods=['POST'])
def upload_image():
    global device_token
    global training_condition
    if 'image' not in request.files:
        return jsonify({"message": "No image file in request"}), 400

    image = request.files['image']

    if image.filename == '':
        return jsonify({"message": "No selected file"}), 400

    if image and allowed_file(image.filename):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = secure_filename(f"{timestamp}_{image.filename}")
            filepath = os.path.join(DOWNLOAD_FOLDER, filename)

            # Save the file
            image.save(filepath)

            # Return the file URL (modify IP to match your Pi IP)
            image_url = f"http://192.168.100.152:5001/downloads/{filename}"
            training_condition=True
            return jsonify({
                "message": "Image uploaded successfully",
                "image_url": image_url,
                "filename": filename
            }), 200

        except Exception as e:
            return jsonify({"message": "Failed to upload image", "error": str(e)}), 500

    return jsonify({"message": "Invalid file type"}), 400

@app.route('/files/<folder>/<filename>')
def get_file(folder, filename):
    if folder == 'uploads':
        return send_from_directory(UPLOAD_FOLDER, filename)
    elif folder == 'downloads':
        return send_from_directory(DOWN_FOLDER, filename)
    else:
        return jsonify({"message": "Invalid folder name"}), 400
@app.route('/save-token', methods=['POST'])
def save_token():
    global device_token
    data = request.get_json()
    device_token = data.get('token')
    print(f"Received device token: {device_token}")
    return "Token received", 200

# --- Registration, Update, Login endpoints (unchanged from your code) ---
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "register_log.json")

def load_logs():
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, 'r') as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def save_logs(logs):
    with open(LOG_FILE, 'w') as f:
        json.dump(logs, f, indent=4)

@app.route('/register', methods=['POST'])
def register():
    logs = load_logs()
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    role = data.get('role')
    if any(log.get("username") == username for log in logs):
        return jsonify({"status": "fail", "message": "Username already exists"}), 409
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "username": username,
        "password": password,
        "role": role
    }
    logs.append(log_entry)
    save_logs(logs)
    return jsonify({"status": "success", "message": "User registered successfully"}), 200

@app.route('/update', methods=['PUT'])
def update_user():
    logs = load_logs()
    data = request.get_json()
    old_username = data.get('old_username')
    old_password = data.get('old_password')
    new_username = data.get('new_username')
    new_password = data.get('new_password')
    updated = False
    for log in logs:
        if log['username'] == old_username and log['password'] == old_password:
            if new_username:
                log['username'] = new_username
            if new_password:
                log['password'] = new_password
            log['timestamp'] = datetime.now().isoformat()
            updated = True
            break
    if updated:
        save_logs(logs)
        return jsonify({"status": "success", "message": "User updated successfully"}), 200
    else:
        return jsonify({"status": "error", "message": "User not found"}), 404

@app.route('/update_role', methods=['PUT'])
def update_role():
    logs = load_logs()
    data = request.get_json()
    username = data.get('username')
    new_role = data.get('new_role')
    updated = False
    for log in logs:
        if log['username'] == username:
            log['role'] = new_role
            log['timestamp'] = datetime.now().isoformat()
            updated = True
            break
    if updated:
        save_logs(logs)
        return jsonify({"status": "success", "message": "Role updated successfully"}), 200
    else:
        return jsonify({"status": "error", "message": "User not found"}), 404

@app.route('/login', methods=['POST'])
def login():
    logs = load_logs()
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    for log in logs:
        if log['username'] == username and log['password'] == password:
            return jsonify({
                "status": "success",
                "message": "Login successful",
                "role": log['role'],
                "username": log['username']
            }), 200
    return jsonify({"status": "error", "message": "Invalid credentials"}), 401

@app.route('/send_notification', methods=['POST'])
def send_notification():
    data = request.get_json()
    if not data or 'filename' not in data:
        return jsonify({"status": "Missing 'filename' in request"}), 400

    full_path = data.get('filename')
    filename = os.path.basename(full_path)  # Extract just "captured_image.jpg"

    image_url = f"http://192.168.100.152:5001/uploads/{filename}"
    if device_token:
        message = messaging.Message(
            notification=messaging.Notification(
                title="New Image from Raspberry Pi",
                body="Tap to view the image"
            ),
            data={"image_url": image_url},
            token=device_token
        )
        try:
            response = messaging.send(message)
            print("Notification sent:", response)
            return jsonify({"status": "Notification sent", "image_url": image_url}), 200
        except Exception as e:
            print("FCM Error:", e)
            return jsonify({"status": "FCM send error", "error": str(e)}), 500
    else:
        return jsonify({"status": "No device token available"}), 400

def run_flask_server():
    app.run(host='0.0.0.0', port=5001)
def trigger_notification(image_path_0):
        filename = image_path_0
        try:
            response = requests.post("http://192.168.100.152:5001/send_notification", json={"filename": filename})
            print("Main Response:", response.json())
        except Exception as e:
            print("Error sending request to Flask:", e)

# Main Pipeline Process

def main_pipeline():
  while True:
    THRESHOLD_CM = 20  # Treat anything below 10cm as 'beam broken'
    global training_condition, is_training, entry_cond,exit_cond
    if training_condition and not is_training.value:
        print("[MAIN PIPELINE] Training condition met. Starting training...")
        training_thread = Thread(target=main_pipeline)
        training_thread.start()
    if lgpio.gpio_read(h, pir1) == 1 and lgpio.gpio_read(h, pir2) == 0:
        entry_cond=True
        print("entry")
        # image_path_0 = capture_image1()
        image_path_1 = capture_image2()
        print(image_path_1)
        detected_faces_paths = detect_faces_and_get_images(image_path_1)
        for face_path in detected_faces_paths:
          recog_result,recognized_label=recognize_face(face_path, threshold=0.6)  # Pass the path to the recognize function
          if recog_result == 1:
              print("1")
              print(recognized_label)
              print("Relay ON (Open)")
              lgpio.gpio_write(h, relay_pin, 1)
              time.sleep(10)
              dist = get_distance()
              print(f"Distance: {dist:.2f} cm")
              if dist < THRESHOLD_CM:
                print("🚨 Beam Broken! Object detected.")
                print("Relay OFF (Closed)")
                lgpio.gpio_write(h, relay_pin, 0)
                record_attendance(recognized_label)
                entry_cond=False
          else:
              trigger_notification(image_path_1)
              shutil.move(image_path_1, os.path.join(UPLOAD_FOLDER, os.path.basename(image_path_1)))
        time.sleep(3)

    if lgpio.gpio_read(h, pir1) == 0 and lgpio.gpio_read(h, pir2) == 1:
        exit_cond=True
        print("exit")
        image_path_0 = capture_image1()
        print(image_path_0)
        detected_faces_paths = detect_faces_and_get_images(image_path_0)
        for face_path in detected_faces_paths:
          recog_result,recognized_label=recognize_face(face_path, threshold=0.6)  # Pass the path to the recognize function
          if recog_result == 1:
              print("1")
              print(recognized_label)
              print("Relay ON (Open)")
              lgpio.gpio_write(h, relay_pin, 1)
              time.sleep(10)
              dist = get_distance()
              print(f"Distance: {dist:.2f} cm")
              if dist < THRESHOLD_CM:
                print("🚨 Beam Broken! Object detected.")
                print("Relay OFF (Closed)")
                lgpio.gpio_write(h, relay_pin, 0)
                record_attendance(recognized_label)
                exit_cond=False

          else:
              trigger_notification(image_path_0)
              shutil.move(image_path_0, os.path.join(UPLOAD_FOLDER, os.path.basename(image_path_0)))
        time.sleep(3)


if __name__ == "__main__":
    try:
        pipeline_thread = Thread(target=main_pipeline)
        flask_thread = Thread(target=run_flask_server)

        pipeline_thread.start()
        flask_thread.start()

        pipeline_thread.join()
        flask_thread.join()

    except KeyboardInterrupt:
        print("[SHUTDOWN] Terminating threads...")


